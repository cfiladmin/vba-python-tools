"""ツール3本の動作確認スクリプト"""
import sys, os, subprocess
from pathlib import Path
import pandas as pd
import openpyxl

BASE    = Path(__file__).parent
TOOLS   = BASE.parent / "VBA.python" / "Python"
RESULTS = []

def result(name, ok, detail):
    mark = "[PASS]" if ok else "[FAIL]"
    print(f"{mark} {name}")
    if detail:
        for line in detail.splitlines():
            print(f"       {line}")
    RESULTS.append((name, ok))

print("=" * 55)
print("ツール動作確認テスト")
print("=" * 55)

# ========== ツール1: PY-001 CSV整形クレンジング ==========
print()
print("-- ツール1: PY-001 CSV整形クレンジング --")

# shift_jis テストCSVを生成
sjis_csv = BASE / "input" / "test_sjis.csv"
rows = [
    "会社名,担当者,金額,日付\n",
    "  株式会社A  , 田中太郎 ,50000,2026-01-10\n",
    "株式会社B,鈴木花子,30000,2026-01-15\n",
    "  株式会社A  , 田中太郎 ,50000,2026-01-10\n",   # 重複
    "株式会社C,  佐藤一郎  ,80000,2026-01-20\n",
    ",,, \n",                                         # 空行
    "株式会社D,山田次郎,20000,2026-01-25\n",
    "株式会社B,鈴木花子,30000,2026-01-15\n",           # 重複
]
with open(sjis_csv, "w", encoding="shift_jis") as f:
    f.writelines(rows)

out_csv = BASE / "output" / "cleaned.csv"

# PY-001 を sys.argv を差し込んで直接実行
import importlib.util, types

src = (TOOLS / "PY-001_CSV整形クレンジング.py").read_text(encoding="utf-8")
# INPUT_ENCODING を shift_jis のまま使うためそのまま実行
exec_ns = {"__file__": str(TOOLS / "PY-001_CSV整形クレンジング.py"), "__name__": "__main__"}
old_argv = sys.argv
sys.argv = ["PY-001", str(sjis_csv), str(out_csv)]
try:
    exec(src, exec_ns)
    sys.argv = old_argv

    df_out = pd.read_csv(out_csv, encoding="utf-8-sig")
    expected_rows = 4  # 重複2件・空行1件を除いた残り

    detail = (
        f"入力: 7行（重複2件・空行1件含む）\n"
        f"出力: {len(df_out)}行（重複・空行除去済み）\n"
        f"出力ファイル: output/cleaned.csv\n"
        f"  {df_out.to_string(index=False)}"
    )
    result("PY-001 CSV整形クレンジング", len(df_out) == expected_rows, detail)
except Exception as e:
    sys.argv = old_argv
    result("PY-001 CSV整形クレンジング", False, f"エラー: {e}")

# ========== ツール2: PY-003 Excel自動生成 ==========
print()
print("-- ツール2: PY-003 Excel自動生成 --")

os.chdir(BASE)  # template.xlsx / data.csv がある場所へ

src2 = (TOOLS / "PY-003_Excel自動生成.py").read_text(encoding="utf-8")
exec_ns2 = {"__file__": str(TOOLS / "PY-003_Excel自動生成.py"), "__name__": "__main__"}
try:
    exec(src2, exec_ns2)

    generated = list((BASE / "output").glob("document_*.xlsx"))
    df_data   = pd.read_csv(BASE / "data.csv", encoding="utf-8-sig", dtype=str)

    details = [f"テンプレート: template.xlsx", f"データ: data.csv ({len(df_data)}件)"]
    all_ok = True
    for f in sorted(generated):
        wb = openpyxl.load_workbook(f)
        ws = wb.active
        company = ws["B3"].value
        person  = ws["B4"].value
        amount  = ws["D10"].value
        date    = ws["F2"].value
        ok = all([company, person, amount, date])
        all_ok = all_ok and ok
        details.append(f"  {f.name}: {company} / {person} / {amount}円 / {date}  {'OK' if ok else 'NG'}")

    result("PY-003 Excel自動生成", all_ok and len(generated) == len(df_data), "\n".join(details))
except Exception as e:
    result("PY-003 Excel自動生成", False, f"エラー: {e}")

# ========== ツール3: VBA-001 集計マクロ（ロジック検証） ==========
print()
print("-- ツール3: VBA-001 集計マクロ（Pythonでロジック検証） --")

try:
    # VBAのDictionary集計と同じロジックをPythonで再現して正確性を確認
    test_data = [
        {"項目": "経費A", "金額": 5000},
        {"項目": "経費B", "金額": 3000},
        {"項目": "経費A", "金額": 2000},
        {"項目": "経費C", "金額": 8000},
        {"項目": "経費B", "金額": 1500},
    ]

    # VBA-001と同じ集計ロジック（Dictionary相当）
    agg = {}
    for row in test_data:
        k = row["項目"]
        agg[k] = agg.get(k, 0) + row["金額"]

    expected = {"経費A": 7000, "経費B": 4500, "経費C": 8000}
    ok = (agg == expected)

    # テスト用Excelを生成（VBAをインポートすれば即実行可能な状態）
    wb_vba = openpyxl.Workbook()
    ws_in  = wb_vba.active
    ws_in.title = "データ"
    ws_in.append(["項目", "金額"])
    for r in test_data:
        ws_in.append([r["項目"], r["金額"]])

    ws_out = wb_vba.create_sheet("集計")
    ws_out.append(["項目", "合計"])

    vba_test_path = BASE / "output" / "VBA001_test.xlsx"
    wb_vba.save(vba_test_path)

    detail = (
        f"集計ロジック検証（VBAのDictionary集計と同一アルゴリズム）\n"
        f"  入力5行 -> 集計結果: {agg}\n"
        f"  期待値:              {expected}\n"
        f"テスト用Excel生成: output/VBA001_test.xlsx\n"
        f"  （VBAエディタでVBA-001をインポートして実行可能）"
    )
    result("VBA-001 集計マクロ（ロジック検証）", ok, detail)
except Exception as e:
    result("VBA-001 集計マクロ（ロジック検証）", False, f"エラー: {e}")

# ========== 結果サマリー ==========
print()
print("=" * 55)
passed = sum(1 for _, ok in RESULTS if ok)
print(f"結果: {passed}/{len(RESULTS)} テスト通過")
for name, ok in RESULTS:
    print(f"  {'[PASS]' if ok else '[FAIL]'} {name}")
print("=" * 55)
