"""テスト用サンプルデータを生成するスクリプト"""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

BASE = Path(__file__).parent
(BASE / "input").mkdir(exist_ok=True)
(BASE / "output").mkdir(exist_ok=True)

# ===== PY-001 用: 汚いCSV（空白・重複・混在あり） =====
rows = [
    ["会社名", "担当者", "金額", "日付"],
    ["  株式会社A  ", " 田中太郎 ", "50000", "2026-01-10"],
    ["株式会社B", "鈴木花子", "30000", "2026-01-15"],
    ["  株式会社A  ", " 田中太郎 ", "50000", "2026-01-10"],   # 重複行
    ["株式会社C", "  佐藤一郎  ", "80000", "2026-01-20"],
    ["", "", "", ""],                                           # 空行
    ["株式会社D", "山田次郎", "20000", "2026-01-25"],
    ["株式会社B", "鈴木花子", "30000", "2026-01-15"],           # 重複行
]

# UTF-8で書き込み（PY-001をUTF-8モードでテスト）
with open(BASE / "input" / "dirty_data.csv", "w", encoding="utf-8-sig", newline="") as f:
    for row in rows:
        f.write(",".join(row) + "\n")
print("[OK] PY-001用テストCSV生成: input/dirty_data.csv")
print(f"  行数: {len(rows)-1}行（重複2件・空行1件含む）")

# ===== PY-003 用: Excelテンプレート =====
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "請求書"

# ヘッダー装飾
header_fill = PatternFill("solid", fgColor="1A2744")
header_font = Font(bold=True, color="FFFFFF", size=12)

# タイトル
ws.merge_cells("A1:F1")
ws["A1"] = "御請求書"
ws["A1"].font = Font(bold=True, size=18, color="1A2744")
ws["A1"].alignment = Alignment(horizontal="center")

# 宛先・日付
ws["A3"] = "宛先："
ws["B3"] = ""            # ← セル B3 が MAPPING["会社名"]
ws["A4"] = "担当者："
ws["B4"] = ""            # ← セル B4 が MAPPING["担当者"]
ws["E2"] = "発行日："
ws["F2"] = ""            # ← セル F2 が MAPPING["日付"]

# 金額欄
ws["C10"] = "請求金額（税込）"
ws["C10"].font = Font(bold=True)
ws["D10"] = ""           # ← セル D10 が MAPPING["金額"]
ws["D10"].number_format = '#,##0"円"'

# 列幅
ws.column_dimensions["A"].width = 12
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 20
ws.column_dimensions["D"].width = 15
ws.column_dimensions["E"].width = 12
ws.column_dimensions["F"].width = 14

wb.save(BASE / "template.xlsx")
print("[OK] PY-003用テンプレートExcel生成: template.xlsx")

# ===== PY-003 用: データCSV =====
df = pd.DataFrame([
    {"会社名": "株式会社アルファ",   "担当者": "田中太郎", "金額": "110000", "日付": "2026-05-23"},
    {"会社名": "有限会社ベータ商事", "担当者": "鈴木花子", "金額": "55000",  "日付": "2026-05-23"},
    {"会社名": "合同会社ガンマ",     "担当者": "佐藤次郎", "金額": "82500",  "日付": "2026-05-23"},
])
df.to_csv(BASE / "data.csv", index=False, encoding="utf-8-sig")
print("[OK] PY-003用データCSV生成: data.csv")
print(f"  件数: {len(df)}件")
print()
print("テストデータ準備完了")
